import os
import csv
import json
import yaml
import toml
import pickle
import unittest
import tempfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

from shiftify import JSON


def read_multiline_json_objects(file_path):
    """
    Helper function to reassemble and read pretty-printed (multi-line)
    JSON objects from a file that writes one JSON object (with indent) per record.
    Assumes that each JSON object ends on a line containing only a closing curly brace.
    """
    objects = []
    with open(file_path, 'r', encoding='utf-8') as f:
        current_lines = []
        for line in f:
            # Skip blank lines.
            if not line.strip():
                continue
            current_lines.append(line)
            if line.strip() == "}":
                try:
                    obj = json.loads("".join(current_lines))
                    objects.append(obj)
                    current_lines = []
                except json.JSONDecodeError:
                    # Continue accumulating lines if object is not complete.
                    pass
        # Try to decode any remaining lines.
        if current_lines:
            try:
                obj = json.loads("".join(current_lines))
                objects.append(obj)
            except json.JSONDecodeError:
                pass
    return objects


class TestJSONConvertor(unittest.TestCase):
    def setUp(self):
        # Create a temporary directory for test files.
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_dir = self.temp_dir.name

        # Prepare a sample JSON file.
        # The JSON file should be a top-level array of objects.
        self.json_file_path = os.path.join(self.test_dir, "test.json")
        self.sample_data = [
            {"name": "Alice", "age": 30, "city": "New York"},
            {"name": "Bob", "age": 25, "city": "Los Angeles"},
            {"name": "Charlie", "age": 35, "city": "Chicago"},
        ]
        # Write the JSON array.
        with open(self.json_file_path, 'w', encoding='utf-8') as f:
            json.dump(self.sample_data, f)

        # Initialize our JSON converter.
        self.converter = JSON(self.json_file_path)

        # Prepare an expected CSV version (note: CSV always yields string values).
        self.expected_csv = [
            {"name": "Alice", "age": "30", "city": "New York"},
            {"name": "Bob", "age": "25", "city": "Los Angeles"},
            {"name": "Charlie", "age": "35", "city": "Chicago"},
        ]

    def tearDown(self):
        # Cleanup the temporary directory.
        self.temp_dir.cleanup()

    def test_to_csv(self):
        csv_file = os.path.join(self.test_dir, "output.csv")
        self.converter.to_csv(csv_file)
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        self.assertEqual(rows, self.expected_csv)

    def test_to_sql(self):
        sql_file = os.path.join(self.test_dir, "output.sql")
        self.converter.to_sql(sql_file_path=sql_file, table_name="people")
        with open(sql_file, 'r', encoding='utf-8') as f:
            content = f.read()
        # Split content into individual SQL statements.
        statements = [s for s in content.splitlines() if s.strip()]
        self.assertEqual(len(statements), len(self.sample_data))
        # For each expected row, verify that at least one statement contains the row's data.
        for expected in self.sample_data:
            found = any(
                expected["name"] in stmt and
                str(expected["age"]) in stmt and
                expected["city"] in stmt
                for stmt in statements
            )
            self.assertTrue(found, f"Expected row {expected} not found in any SQL statement.")

    def test_to_html(self):
        html_file = os.path.join(self.test_dir, "output.html")
        self.converter.to_html(html_file)
        with open(html_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
        self.assertIn("<table", html_content)
        self.assertIn("<th>name</th>", html_content)
        self.assertIn("<td>Alice</td>", html_content)

    def test_to_text(self):
        text_file = os.path.join(self.test_dir, "output.txt")
        self.converter.to_text(text_file)
        objects = read_multiline_json_objects(text_file)
        self.assertEqual(len(objects), 3)
        for obj, expected in zip(objects, self.sample_data):
            self.assertEqual(obj, expected)

    def test_to_yaml(self):
        yaml_file = os.path.join(self.test_dir, "output.yaml")
        self.converter.to_yaml(yaml_file)
        with open(yaml_file, 'r', encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)
        self.assertIsInstance(yaml_data, list)
        self.assertEqual(yaml_data, self.sample_data)

        yaml_file_key = os.path.join(self.test_dir, "output_key.yaml")
        self.converter.to_yaml(yaml_file_key, key="record")
        with open(yaml_file_key, 'r', encoding='utf-8') as f:
            yaml_data_key = yaml.safe_load(f)
        self.assertIsInstance(yaml_data_key, list)
        for item in yaml_data_key:
            self.assertIn("record", item)
            # Check that the value under 'record' is one of the sample objects.
            self.assertIn(item["record"], self.sample_data)

    def test_to_xml(self):
        xml_file = os.path.join(self.test_dir, "output.xml")
        self.converter.to_xml(xml_file, root_element="People")
        tree = ET.parse(xml_file)
        root = tree.getroot()
        self.assertEqual(root.tag, "People")
        items = root.findall("Item")
        self.assertEqual(len(items), 3)
        for item in items:
            for key in ["name", "age", "city"]:
                element = item.find(key)
                self.assertIsNotNone(element)

    def test_to_tsv(self):
        tsv_file = os.path.join(self.test_dir, "output.tsv")
        self.converter.to_tsv(tsv_file)
        with open(tsv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            rows = list(reader)
        self.assertEqual(rows[0], ["name", "age", "city"])
        self.assertEqual(len(rows), 4)

    def test_to_toml(self):
        toml_file = os.path.join(self.test_dir, "output.toml")
        self.converter.to_toml(toml_file)
        with open(toml_file, 'r', encoding='utf-8') as f:
            toml_data = toml.load(f)
        for i in range(3):
            key = f"item_{i}"
            self.assertIn(key, toml_data)
            self.assertEqual(toml_data[key], self.sample_data[i])

    def test_to_ndjson(self):
        ndjson_file = os.path.join(self.test_dir, "output.ndjson")
        self.converter.to_ndjson(ndjson_file)
        with open(ndjson_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        self.assertEqual(len(lines), 3)
        for line, expected in zip(lines, self.sample_data):
            data = json.loads(line)
            self.assertEqual(data, expected)

    def test_to_excel(self):
        excel_file = os.path.join(self.test_dir, "output.xlsx")
        self.converter.to_excel(excel_file, sheet_name="Data")
        wb = load_workbook(excel_file)
        sheet = wb["Data"]
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header, ["name", "age", "city"])
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)
        for row, expected in zip(rows, self.sample_data):
            self.assertEqual(row, (expected["name"], expected["age"], expected["city"]))

    def test_to_markdown_table(self):
        md_file = os.path.join(self.test_dir, "output.md")
        self.converter.to_markdown_table(md_file)
        with open(md_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        self.assertTrue(lines[0].startswith("|"))
        self.assertTrue(lines[1].strip().startswith("|"))
        self.assertEqual(len(lines), 1 + 1 + 3)

    def test_to_binary(self):
        binary_file = os.path.join(self.test_dir, "output.bin")
        self.converter.to_binary(binary_file)
        unpickled_items = []
        with open(binary_file, 'rb') as f:
            while True:
                length_bytes = f.read(4)
                if not length_bytes:
                    break
                length = int.from_bytes(length_bytes, byteorder='big')
                binary_data = f.read(length)
                unpickled_items.append(pickle.loads(binary_data))
        self.assertEqual(unpickled_items, self.sample_data)

    def test_to_ini(self):
        ini_file = os.path.join(self.test_dir, "output.ini")
        self.converter.to_ini(ini_file, section_prefix="Person")
        with open(ini_file, 'r', encoding='utf-8') as f:
            content = f.read()
        self.assertIn("[Person_1]", content)
        self.assertIn("name = Alice", content)
        self.assertIn("[Person_3]", content)
        self.assertIn("city = Chicago", content)

    # Negative / edge case tests.
    def test_non_existent_file(self):
        non_existent = os.path.join(self.test_dir, "does_not_exist.json")
        with self.assertRaises(FileNotFoundError):
            JSON(non_existent)

    def test_empty_file(self):
        empty_file = os.path.join(self.test_dir, "empty.json")
        with open(empty_file, 'w', encoding='utf-8') as f:
            pass
        with self.assertRaises(ValueError):
            JSON(empty_file)


if __name__ == '__main__':
    unittest.main()
