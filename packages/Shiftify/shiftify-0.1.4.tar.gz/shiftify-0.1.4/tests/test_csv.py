
import os
import csv
import json
import yaml
import toml
import pickle
import unittest
import tempfile
import xml.etree.ElementTree as ET
from xml.dom.minidom import parseString
from openpyxl import load_workbook

# Import the CSV class from your module.
# For example, if your module is named csv_convertor.py:
from shiftify import CSV

def read_multiline_json_objects(file_path):
    """
    Helper function to read a file containing multiple pretty-printed JSON objects.
    Each JSON object is assumed to end with a line that contains only a closing curly brace ('}').
    Returns a list of deserialized JSON objects.
    """
    objects = []
    with open(file_path, 'r', encoding='utf-8') as f:
        current = ""
        for line in f:
            # Skip blank lines.
            if not line.strip():
                continue
            current += line
            # If the line (stripped of whitespace) is a closing brace, assume the object is complete.
            if line.strip() == "}":
                try:
                    obj = json.loads(current)
                    objects.append(obj)
                    current = ""
                except json.JSONDecodeError:
                    # If not complete yet, continue accumulating.
                    pass
        # In case any remaining text forms a complete JSON object:
        if current.strip():
            try:
                obj = json.loads(current)
                objects.append(obj)
            except json.JSONDecodeError:
                pass
    return objects

class TestCSV(unittest.TestCase):
    def setUp(self):
        # Create a temporary directory for test files.
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_dir = self.temp_dir.name

        # Prepare a sample CSV file.
        self.csv_file_path = os.path.join(self.test_dir, "test.csv")
        sample_csv = (
            "name,age,city\n"
            "Alice,30,New York\n"
            "Bob,25,Los Angeles\n"
            "Charlie,35,Chicago\n"
        )
        with open(self.csv_file_path, 'w', encoding='utf-8') as f:
            f.write(sample_csv)

        # Initialize our CSV converter.
        self.converter = CSV(self.csv_file_path)

        # Expected data (note: CSV data are strings)
        self.expected_data = [
            {"name": "Alice", "age": "30", "city": "New York"},
            {"name": "Bob", "age": "25", "city": "Los Angeles"},
            {"name": "Charlie", "age": "35", "city": "Chicago"},
        ]

    def tearDown(self):
        # Cleanup the temporary directory.
        self.temp_dir.cleanup()

    def test_to_json(self):
        json_file = os.path.join(self.test_dir, "output.json")
        self.converter.to_json(json_file)

        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.assertEqual(data, self.expected_data)

    def test_to_sql(self):
        sql_file = os.path.join(self.test_dir, "output.sql")
        self.converter.to_sql(sql_file_path=sql_file, table_name="people")
        with open(sql_file, 'r', encoding='utf-8') as f:
            content = f.read()
        # Check that there are three INSERT statements
        statements = [s for s in content.splitlines() if s.strip()]
        self.assertEqual(len(statements), 3)
        for stmt in statements:
            self.assertTrue(stmt.startswith("INSERT INTO people"))
            self.assertIn("VALUES", stmt)

    def test_to_html(self):
        html_file = os.path.join(self.test_dir, "output.html")
        self.converter.to_html(html_file)
        with open(html_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
        # Check for basic HTML structure.
        self.assertIn("<table", html_content)
        self.assertIn("<th>name</th>", html_content)
        self.assertIn("<td>Alice</td>", html_content)

    def test_to_text(self):
        text_file = os.path.join(self.test_dir, "output.txt")
        self.converter.to_text(text_file)
        # Use helper function to reassemble each multi-line JSON object.
        objects = read_multiline_json_objects(text_file)
        self.assertEqual(len(objects), 3)
        for obj, expected in zip(objects, self.expected_data):
            self.assertEqual(obj, expected)
            
    def test_to_yaml(self):
        yaml_file = os.path.join(self.test_dir, "output.yaml")
        self.converter.to_yaml(yaml_file)
        with open(yaml_file, 'r', encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)
        # Without a key, we expect a list of dictionaries.
        self.assertIsInstance(yaml_data, list)
        self.assertEqual(yaml_data, self.expected_data)

        # Test with a wrapping key.
        yaml_file_key = os.path.join(self.test_dir, "output_key.yaml")
        self.converter.to_yaml(yaml_file_key, key="record")
        with open(yaml_file_key, 'r', encoding='utf-8') as f:
            yaml_data_key = yaml.safe_load(f)
        self.assertIsInstance(yaml_data_key, list)
        for item in yaml_data_key:
            self.assertIn("record", item)

    def test_to_xml(self):
        xml_file = os.path.join(self.test_dir, "output.xml")
        self.converter.to_xml(xml_file, root_element="People")
        # Parse the XML and check structure.
        tree = ET.parse(xml_file)
        root = tree.getroot()
        self.assertEqual(root.tag, "People")
        items = root.findall("Item")
        self.assertEqual(len(items), 3)
        # Check that each <Item> contains expected subelements.
        for item in items:
            for key in ["name", "age", "city"]:
                self.assertIsNotNone(item.find(key))

    def test_to_tsv(self):
        tsv_file = os.path.join(self.test_dir, "output.tsv")
        self.converter.to_tsv(tsv_file)
        with open(tsv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            rows = list(reader)
        # Check header and number of rows.
        self.assertEqual(rows[0], ["name", "age", "city"])
        self.assertEqual(len(rows), 4)

    def test_to_toml(self):
        toml_file = os.path.join(self.test_dir, "output.toml")
        self.converter.to_toml(toml_file)
        with open(toml_file, 'r', encoding='utf-8') as f:
            toml_data = toml.load(f)
        # Expect keys item_0, item_1, item_2 in the output.
        for i in range(3):
            key = f"item_{i}"
            self.assertIn(key, toml_data)
            self.assertEqual(toml_data[key], self.expected_data[i])

    def test_to_ndjson(self):
        ndjson_file = os.path.join(self.test_dir, "output.ndjson")
        self.converter.to_ndjson(ndjson_file)
        with open(ndjson_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        self.assertEqual(len(lines), 3)
        for line, expected in zip(lines, self.expected_data):
            data = json.loads(line)
            self.assertEqual(data, expected)

    def test_to_excel(self):
        excel_file = os.path.join(self.test_dir, "output.xlsx")
        self.converter.to_excel(excel_file, sheet_name="Data")
        wb = load_workbook(excel_file)
        sheet = wb["Data"]
        # Check that header row is present.
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header, ["name", "age", "city"])
        # Check that there are three more rows.
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)
        for row, expected in zip(rows, self.expected_data):
            # Convert expected values (all strings) to tuple.
            expected_tuple = (expected["name"], expected["age"], expected["city"])
            self.assertEqual(row, expected_tuple)

    def test_to_markdown_table(self):
        md_file = os.path.join(self.test_dir, "output.md")
        self.converter.to_markdown_table(md_file)
        with open(md_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        # The first two lines are header and separator.
        self.assertTrue(lines[0].startswith("|"))
        self.assertTrue(lines[1].strip().startswith("|"))
        # There should be a header, a separator, and one row per CSV record.
        self.assertEqual(len(lines), 1 + 1 + 3)

    def test_to_binary(self):
        binary_file = os.path.join(self.test_dir, "output.bin")
        self.converter.to_binary(binary_file)
        unpickled_rows = []
        with open(binary_file, 'rb') as f:
            while True:
                length_bytes = f.read(4)
                if not length_bytes:
                    break
                length = int.from_bytes(length_bytes, byteorder='big')
                binary_data = f.read(length)
                unpickled_rows.append(pickle.loads(binary_data))
        self.assertEqual(unpickled_rows, self.expected_data)

    def test_to_ini(self):
        ini_file = os.path.join(self.test_dir, "output.ini")
        self.converter.to_ini(ini_file, section_prefix="Person")
        with open(ini_file, 'r', encoding='utf-8') as f:
            content = f.read()
        # Check for expected section headers and key=value pairs.
        self.assertIn("[Person_1]", content)
        self.assertIn("name = Alice", content)
        self.assertIn("[Person_3]", content)
        self.assertIn("city = Chicago", content)

    # Negative and edge case tests
    def test_non_existent_file(self):
        non_existent = os.path.join(self.test_dir, "does_not_exist.csv")
        with self.assertRaises(FileNotFoundError):
            CSV(non_existent)

    def test_empty_file(self):
        empty_csv = os.path.join(self.test_dir, "empty.csv")
        # Create an empty file.
        with open(empty_csv, 'w', encoding='utf-8') as f:
            pass
        with self.assertRaises(ValueError):
            CSV(empty_csv)

if __name__ == '__main__':
    unittest.main()
