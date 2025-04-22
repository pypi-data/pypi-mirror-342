# -*- coding: UTF-8 -*-
# python3

import os
import sys
import json
import requests
from io import BytesIO
from openpyxl import load_workbook

# PC SDK build tool

from devolib import DynamicObject
from devolib.util_log import LOG_D, LOG_E, LOG_W, LOG_I
from devolib.util_os import get_env_var, current_dir
from devolib.util_str import starts_with, has, replace
from devolib.util_argparse import typeparse_str2bool
from devolib.util_httpc import GET_JSON
from devolib.util_crypt import sim_cipher_decrypt, aes_encrypt_without_b64, aes_decrypt_without_b64
from devolib.util_fs import path_join_one, write_bytes_to_file, path_exists, read_bytes_of_file, write_file, touch_dir, copy_files, copy_dir, remove_files, remove_file
from devolib.util_json import json_to_str, json_from_str
from devolib.consts import ENV_DEV, ENV_TEST, ENV_PROD, PLAT_ADR, PLAT_WIN, PLAT_IOS

# MARK: Consts

CIPHER_FOR_CIPHER_BYTES = [0xc7, 0xc4, 0xc5, 0xda, 0xcb, 0xcf, 0xcc, 0xcd, 0xc2, 0xc3, 0xc0, 0xc4, 0xc5, 0xda, 0xdb, 0xd8]
CIPHER_FOR_CIPHER_SALT = 0xAA
CIPHER_FOR_CIPHER_IV = [0x9b, 0x98, 0xcb, 0xcb, 0xec, 0xee, 0xf9, 0xeb, 0xc1, 0xcb, 0xc7, 0xcc, 0xce, 0xd9, 0xcb, 0x9b]

# MARK: Utils

def conf_is_wegame(conf_json): # conf = conf_json["data"]
    conf = conf_json["data"]
    return conf["store_type"] == "wegame"

def conf_is_steam(conf_json):
    conf = conf_json["data"]
    return conf["store_type"] == "steam"

def conf_is_offcial(conf_json):
    conf = conf_json["data"]
    return conf["store_type"] == "offcial"

def find_streaming_assets(build_dir):
    """
    查找指定目录下的 StreamingAssets 子目录路径。
    如果找到，返回其路径；如果未找到，则返回 None。
    """
    if not os.path.isdir(build_dir):
        raise ValueError(f"Provided {build_dir} is not valid dir")
    
    for root, dirs, _ in os.walk(build_dir):
        if "StreamingAssets" in dirs:
            return os.path.join(root, "StreamingAssets")
    
    return None

# MARK: Data retrieve

def parse_excel_data(file_stream, allowed_headers=None, header_mapping=None):
    # 加载 Excel 文件
    workbook = load_workbook(file_stream, data_only=True)
    sheet = workbook.active

    # 获取表头
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # 应用表头映射
    if header_mapping:
        mapped_headers = [header_mapping.get(header, header) for header in headers]
    else:
        mapped_headers = headers

    # 筛选表头
    if allowed_headers is not None:
        header_indices = [i for i, header in enumerate(mapped_headers) if header in allowed_headers]
        filtered_headers = [mapped_headers[i] for i in header_indices]
    else:
        header_indices = list(range(len(mapped_headers)))
        filtered_headers = mapped_headers

    # 解析数据行
    data = []
    for row in sheet.iter_rows(min_row=2):
        row_data = {filtered_headers[i]: row[header_indices[i]].value for i in range(len(header_indices))}
        data.append(row_data)

    return data

def try_load_local_excel(allowed_headers=None, header_mapping=None):    
    file_path = os.path.join(current_dir(), 'resources', 'sdk-pipe-settings.xlsx')

    LOG_D(f"loadl excel file_path: {file_path}")

    if path_exists(file_path):
        with open(file_path, 'rb') as f:
            file_stream = BytesIO(f.read())

            return parse_excel_data(file_stream, allowed_headers, header_mapping)
    
    return None
    

def download_and_parse_excel(url, allowed_headers=None, header_mapping=None):
    """
    下载 Excel 文件到内存并解析为字典数组。

    :param url: Excel 文件的 URL。
    :param allowed_headers: 要保留的表头列表。如果为 None，则保留所有表头。
    :return: 包含字典的列表，每个字典表示一行记录。
    """
    # 下载文件到内存
    response = requests.get(url)
    response.raise_for_status()
    file_stream = BytesIO(response.content)

    return parse_excel_data(file_stream, allowed_headers, header_mapping)

def get_pipe_data():
    pipe_settings_link = "http://10.36.4.108:8043/shared/sdk-pipe-settings.xlsx"
    allowed_headers = ["name", "env", "region", "platform", "app_id", "store_type", "pipe", "sdk_build_params", "sdk_build_host", "sdk_build_token"]
    headers_mapping = {
        "名称": "name",
        "环境": "env",
        "地区": "region",
        "对应管线": "pipe"
    }

    local_data = try_load_local_excel(allowed_headers, headers_mapping)
    if local_data is not None:
        LOG_D("load pipe info from local.")
        return local_data
    
    LOG_D("load pipe info from remote.")
    
    # 下载文件到内存
    data = download_and_parse_excel(pipe_settings_link, allowed_headers, headers_mapping)

    LOG_D(f"excel data: {json.dumps(data, indent=4)}")

    return data

# a, b, c = parse_dict_data(
#     dict_list=data,
#     target_value=2,
#     key_field="id",
#     return_fields=["name", "age", "email"]
# )
def parse_dict_data(dict_list, target_value, key_field, return_fields):
    """
    根据目标字段值，从字典数组中找到匹配的字典并返回指定字段值的元组。

    :param dict_list: 字典数组。
    :param target_value: 用于匹配的目标值。
    :param key_field: 用于匹配的字典键。
    :param return_fields: 要返回的字段列表。
    :return: 元组，包含返回字段的值。如果没有匹配，返回 None。
    """
    for record in dict_list:
        LOG_D(f'key_field: {record.get(key_field)}, target_value: {target_value}')
        if record.get(key_field) == target_value:
            return tuple(record.get(field) for field in return_fields)
    return None

def get_conf_data(params_str, host, token):
    if params_str != None:
        param_arr = params_str.split("-") # official-pc-10001
        host = host
        res_json = GET_JSON(
            host=f'https://{host}', 
            path='/pconf/pack', 
            query=f"app_id={param_arr[2]}&store_type={param_arr[0]}&platform={param_arr[1]}",
            headers={
                'Authorization': token
            })
        
        if res_json is None:
            raise Exception(f'get conf data failed.')

        code = res_json['code']
        if code != 200:
            raise Exception(f'get conf data failed, code: {code}')

        return res_json
    else:
        LOG_E('host empty')

        return None

# MARK: Build Stages

def stage_get_conf(params, context):
    conf_json = get_conf_data(params, context.host, context.token)

    LOG_W(f"[STAGE] conf json: {conf_json}")

    return conf_json

def stage_handle_files(conf_json, origin_dir, target_dir, is_strict_mode):
    # copy data dir from dll_origin_dir to dll_target_dir
    LOG_D(f"copying `data`")
    origin_dir_data = f"{origin_dir}/data"
    target_dir_data = f"{target_dir}/data"
    if not path_exists(origin_dir_data):
        raise Exception("`data` not found!")
    
    copy_dir(src=origin_dir_data, dst=target_dir_data)
    
    # copy locales
    LOG_D(f"copying `locales`")
    origin_dir_locales = f"{origin_dir}/locales"
    target_dir_locales = f"{target_dir}/locales"
    if not path_exists(origin_dir_locales):
        raise Exception("`locales` not found!")
    
    copy_dir(src=origin_dir_locales, dst=target_dir_locales)

    # copy resources
    LOG_D(f"copying `resources`")
    origin_dir_resources = f"{origin_dir}/resources"
    target_dir_resources = f"{target_dir}/resources"
    if not path_exists(origin_dir_resources):
        raise Exception("`resources` not found!")
    
    copy_dir(src=origin_dir_resources, dst=target_dir_resources)

    # copy files
    LOG_D(f"copying `files`")
    extra_files = [
        f"{origin_dir}/chrome_100_percent.pak",
        f"{origin_dir}/chrome_200_percent.pak",
        f"{origin_dir}/icudtl.dat",
        f"{origin_dir}/limpcbrowser.exe",
        f"{origin_dir}/limpcbrowserex.exe",
        f"{origin_dir}/resources.pak",
        f"{origin_dir}/snapshot_blob.bin",
        f"{origin_dir}/v8_context_snapshot.bin",
        f"{origin_dir}/vk_swiftshader_icd.json"
    ]
    copy_files(file_paths=extra_files, output_dir=target_dir)

    # custom actions for different channels
    LOG_D(f"custom actions")

    if not conf_is_wegame(conf_json) and is_strict_mode:
        LOG_D(f"not wegame: deleting `rail_api64.dll`")
        to_delete_files = [
            f"{target_dir}/rail_api64.dll"
        ]
        remove_files(file_paths=to_delete_files)

    if not conf_is_steam(conf_json) and is_strict_mode:
        LOG_D(f"not steam: deleting `steam_api64.dll`")
        to_delete_files = [
            f"{target_dir}/steam_api64.dll"
        ]
        remove_files(file_paths=to_delete_files)

def stage_save_conf(conf_json, target_dir, need_encrypt, file_name): # 加密配置数据，保存在指定目录
    global CIPHER_FOR_CIPHER_BYTES, CIPHER_FOR_CIPHER_SALT

    LOG_D(f"target_dir: {target_dir}")

    conf_file_path = path_join_one(target_dir, file_name)
    LOG_D(f"conf_file_path: {conf_file_path}")

    conf_json_str = json_to_str(conf_json)
    LOG_D(f"conf_json_str: {conf_json_str}")

    remove_file(conf_file_path)

    if need_encrypt:
        cipher_decrypted = sim_cipher_decrypt(cipher_bytes=CIPHER_FOR_CIPHER_BYTES, salt=CIPHER_FOR_CIPHER_SALT)
        iv_decrypted = sim_cipher_decrypt(cipher_bytes=CIPHER_FOR_CIPHER_IV, salt=CIPHER_FOR_CIPHER_SALT)

        conf_str_encrypted = aes_encrypt_without_b64(conf_json_str, cipher_decrypted, iv_decrypted)
        write_bytes_to_file(conf_file_path, conf_str_encrypted)
    else:
        write_file(conf_file_path, conf_json_str)

    if path_exists(conf_file_path):
        LOG_I(f"conf path exists: {conf_file_path}")
    else:
        LOG_E(f"conf path not exists")

# MARK: Command Handle
def cmd_handle_tools_decrypt(args):
    if args.path is not None:
        file_path = args.path
        decrypted_file_path = f"{file_path}.decrypted"

        LOG_D(f"file_path: {file_path}")
        LOG_D(f"decrypted_file_path: {decrypted_file_path}")

        cipher_decrypted = sim_cipher_decrypt(cipher_bytes=CIPHER_FOR_CIPHER_BYTES, salt=CIPHER_FOR_CIPHER_SALT)
        iv_decrypted = sim_cipher_decrypt(cipher_bytes=CIPHER_FOR_CIPHER_IV, salt=CIPHER_FOR_CIPHER_SALT)

        encrypted_bytes = read_bytes_of_file(file_path)
        decrypted_bytes = aes_decrypt_without_b64(encrypted_bytes, cipher_decrypted, iv_decrypted)
        write_bytes_to_file(decrypted_file_path, decrypted_bytes)
    else:
        LOG_E(f"none file is processed.")

    sys.exit()

def cmd_handle_postbuild(args):
    LOG_D(f'json: {args.json}')

    LOG_D(f'params: {args.params}')
    LOG_D(f'hosts: {args.hosts}')
    LOG_D(f'token: {args.token}')

    LOG_D(f'origin: {args.origin}')
    LOG_D(f'target: {args.target}')
    LOG_D(f'encrypt: {args.encrypt}')
    LOG_D(f'strict: {args.strict}')

    json_str = args.json
    if json_str is None:
        json_str = get_env_var("SDK_BUILD_JSON")
    
    sdk_hosts = None
    sdk_params = None
    sdk_token = None

    json = json_from_str(json_str)
    if 'pipe' in json:
        pipe_settings = get_pipe_data()
        pipe_name = json["pipe"]
        LOG_D(f'pipe: {pipe_name}')

        if 'QB_' in pipe_name:
            key_field = 'name'
        elif '_' in pipe_name:
            key_field = 'sdk_build_params'
        else:
            key_field = 'pipe'

        if pipe_settings is not None:
            store_type, platform, app_id, sdk_build_host, sdk_build_token = parse_dict_data(
                dict_list=pipe_settings,
                target_value=pipe_name,
                key_field=key_field,
                return_fields=["store_type", "platform", "app_id", "sdk_build_host", "sdk_build_token"]
            )
            sdk_params = f"{store_type}-{platform}-{app_id}"
            sdk_hosts = sdk_build_host
            sdk_token = sdk_build_token
    elif 'sdk_params' in json and 'sdk_hosts' in json and 'sdk_token' in json:
        sdk_params = json['sdk_params']
        sdk_hosts = json['sdk_hosts']
        sdk_token = json['sdk_token']
    elif args.params is not None and args.hosts is not None and args.token is not None:
        sdk_params = args.params
        sdk_hosts = args.hosts
        sdk_token = args.token
    else:
        LOG_W("`pipe` is None or `sdk_params/sdk_hosts/sdk_token` is None")
    
    LOG_D(f'sdk_params: {sdk_params}')
    LOG_D(f'sdk_hosts: {sdk_hosts}')
    LOG_D(f'sdk_token: {sdk_token}')  

    if sdk_params is not None and sdk_hosts is not None and sdk_token is not None:
        params_arr = sdk_params.split(",")
        if len(params_arr) == 0:
            raise Exception("params is empty")
        
        for env_key in params_arr:
            context = DynamicObject(host=sdk_hosts, token=sdk_token)

            conf_json = stage_get_conf(env_key, context)

            if has(env_key, PLAT_WIN):
                if args.origin != None and len(args.origin) != 0:
                    stage_handle_files(conf_json, args.origin, args.target, args.strict)

                if args.target == None or len(args.target) == 0:
                    args.target = current_dir()

                # try get env
                env_key = f'{conf_json["data"]["env"]}_{env_key}'
                env_key = replace(env_key, "-", "_")

                if starts_with(sdk_params, env_key) and len(params_arr) > 1:
                    file_name = f"pcsdk.{env_key}.default.json"
                else:
                    file_name = f"pcsdk.{env_key}.json"

            elif has(env_key, PLAT_ADR) or has(env_key, PLAT_IOS):
                args.encrypt = False
                file_name = f"env.json"

                # if args.target is not None: # 当前 target 是构建目标的根目录
                #     args.target = find_streaming_assets(args.target)

                if args.target == None or len(args.target) == 0:
                    args.target = current_dir()

                    LOG_W(f'target use: {args.target}') 
            else:
                raise Exception("Unkown platform")

            touch_dir(args.target) # touch target dir
            stage_save_conf(conf_json, args.target, args.encrypt, file_name)

def cmd_handle_chkenv(args):
    pipe_settings = get_pipe_data()

    checked_params = []

    for setting in pipe_settings:
        key_field = 'sdk_build_params'

        store_type, platform, app_id, sdk_build_host, sdk_build_token = parse_dict_data(
                dict_list=pipe_settings,
                target_value=setting[key_field],
                key_field=key_field,
                return_fields=["store_type", "platform", "app_id", "sdk_build_host", "sdk_build_token"]
            )
        sdk_params = f"{store_type}-{platform}-{app_id}"
        sdk_hosts = sdk_build_host
        sdk_token = sdk_build_token

        context = DynamicObject(host=sdk_hosts, token=sdk_token)
        try:
            conf_json = stage_get_conf(sdk_params, context)
            if conf_json is None:
                # raise Exception("")
                LOG_E(f'error env: {sdk_params}')
            else:
                checked_params.append(sdk_params)
        except:
            LOG_E(f'error env: {sdk_params}')

    if len(checked_params) > 0:
        LOG_D(f'checked params: ')
    for i, p in enumerate(checked_params):
        LOG_D(f'{i+1}: {p}')

# MARK: Command Regist

def cmd_regist(subparsers):
    parser = subparsers.add_parser('sbt.tools.decrypt', help='pc sdk tools for file decryption.')
    parser.add_argument('-p', '--path', type=str, default=None, help='encrypted file path, default search current dir for `pcsdk.json` file.')
    parser.set_defaults(handle=cmd_handle_tools_decrypt)

    parser = subparsers.add_parser('omni.postbuild', help='SDK Build Tool for postbuilding sdk in unity.')
    parser.add_argument('-p', '--params', type=str, default=None, help='store params, E.g: offcial-pc-10001,official-pc-12001')
    parser.add_argument('-ho', '--hosts', type=str, default=None, help='hosts')
    parser.add_argument('-to', '--token', type=str, default=None, help='token')

    parser.add_argument('-j', '--json', type=str, default=None, help='params json string')
    parser.add_argument('-o', '--origin', type=str, default=None, help='origin full path')
    parser.add_argument('-t', '--target', type=str, default=None, help='target full path')
    parser.add_argument('-e', '--encrypt', type=typeparse_str2bool, default=True, help='need encrypt or not')
    parser.add_argument('-s', '--strict', type=typeparse_str2bool, default=True, help='if strict, then file cropping.')
    parser.set_defaults(handle=cmd_handle_postbuild)

    parser = subparsers.add_parser('omni.chkenv', help='SDK Build Tool for check env excel.')
    parser.set_defaults(handle=cmd_handle_chkenv)

# python src/devocli/pcs_postbuild.py
if __name__ == '__main__':
    args = DynamicObject(path="/Users/fallenink/Desktop/Developer/devokay-py/tmp/pcsdk.json")
    cmd_handle_tools_decrypt(args)